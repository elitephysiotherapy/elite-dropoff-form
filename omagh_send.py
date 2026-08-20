"""Send the Omagh launch SMS, in waves, from the tidied send list.

Reads the WORKBOOK (Martin edits it by hand — removing people, fixing numbers),
never the CSV, so manual edits are always what gets sent.

  python omagh_send.py                  # dry run, whole list
  python omagh_send.py --wave 100       # dry run, first 100 unsent
  python omagh_send.py --wave 100 --send

Safety, in order of importance:
  - dry run unless --send is passed
  - every recipient is checked against sent_log BEFORE sending and logged
    IMMEDIATELY after, so a crash mid-wave cannot re-text anyone
  - numbers are re-validated here; the workbook's SMS-ready column is
    hand-editable and has already contained a 10-digit "mobile"
  - MARKETING_SAFE_MODE (default on) reroutes every message to the test phone
  - waves exist because replies arrive in a 2-4 hour burst, not over a week

Sends from config.SMS_SENDER_NUMBER so patients can reply; replies land in
#omagh-replies via /twilio/inbound.
"""

import argparse
import glob
import os
import re
import sys

from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"),
            override=True)

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import config                                    # noqa: E402
from marketing import sent_log, templates, twilio_client   # noqa: E402

FLOW = "omagh_launch"
TEMPLATE = "omagh_launch"
ANCHOR = "2026-09-omagh"      # one send per patient for this campaign, ever
PRICE_PER_SEGMENT = 0.042325  # GBP, from Twilio's UK pricing API


def latest_workbook():
    hits = sorted(glob.glob(os.path.expanduser(
        "~/Downloads/Omagh_Launch_List_*.xlsx")))
    if not hits:
        sys.exit("no Omagh_Launch_List_*.xlsx in ~/Downloads")
    return hits[-1]


def valid_uk_mobile(raw):
    """E.164 for a real UK/IE mobile, else "".

    Deliberately re-derived from the digits rather than trusting the
    SMS-ready column, which is hand-editable.
    """
    d = re.sub(r"\D", "", raw or "")
    if d.startswith("44"):
        d = "0" + d[2:]
    elif d.startswith("353"):
        d = "0" + d[3:]
    if len(d) == 11 and d.startswith("07"):
        return "+44" + d[1:]
    if len(d) == 10 and d.startswith("08"):
        return "+353" + d[1:]
    return ""


def truthy(v):
    return v is True or str(v).strip().lower() == "true"


def load_recipients(path):
    ws = load_workbook(path, data_only=True)["Send List"]
    hdr = [c.value for c in ws[1]]
    H = {h: i for i, h in enumerate(hdr) if h}
    good, bad = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(x not in (None, "") for x in row):
            continue
        name = row[H["Name"]] or ""
        if not truthy(row[H["SMS marketing OK"]]):
            continue
        number = valid_uk_mobile(row[H["SMS-ready number"]]
                                 or row[H["Mobile"]])
        first = (row[H["First name"]] or "").strip() or name.split(" ")[0]
        rec = {"pid": str(row[H["Patient ID"]]), "name": name,
               "first": first, "to": number,
               "town": row[H["Town"]] or "",
               "raw": row[H["SMS-ready number"]] or row[H["Mobile"]] or ""}
        (good if number else bad).append(rec)
    return good, bad


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--send", action="store_true", help="actually send")
    ap.add_argument("--wave", type=int, default=0, help="cap this run")
    args = ap.parse_args()

    path = latest_workbook()
    good, bad = load_recipients(path)
    print(f"list: {os.path.basename(path)}")
    print(f"opted in to SMS: {len(good) + len(bad)}")
    if bad:
        print(f"\n{len(bad)} EXCLUDED — not a valid UK/IE mobile:")
        for r in bad:
            print(f"   {r['name']:<24} {r['town']:<14} {r['raw']!r}")

    # Ledger check before anything is sent.
    todo = [r for r in good
            if not sent_log.already_sent(r["pid"], FLOW, ANCHOR,
                                         within_days=3650,
                                         ignore_failed=True)]
    already = len(good) - len(todo)
    if already:
        print(f"\n{already} already texted for this campaign — skipping")

    if args.wave:
        todo = todo[:args.wave]

    sample = templates.render_sms(TEMPLATE, {"first_name": "Ann - Marie"})
    print(f"\nmessage ({len(sample)} chars, 1 segment):\n  {sample}")
    print(f"\nTO SEND NOW: {len(todo)}"
          f"   (~£{len(todo) * PRICE_PER_SEGMENT:.2f})")
    print(f"from: {config.SMS_SENDER_NUMBER}")
    if config.MARKETING_SAFE_MODE:
        print(f"SAFE MODE ON — everything reroutes to "
              f"{config.MARKETING_TEST_PHONE}")

    if not args.send:
        print("\nDRY RUN — nothing sent. Add --send to go live.")
        for r in todo[:5]:
            print(f"   would text {r['name']} ({r['town']}) {r['to']}")
        if len(todo) > 5:
            print(f"   ...and {len(todo) - 5} more")
        return

    if not todo:
        print("\nnothing to send")
        return

    print(f"\nSENDING to {len(todo)}...")
    ok = fail = 0
    for i, r in enumerate(todo, 1):
        body = templates.render_sms(TEMPLATE, {"first_name": r["first"]})
        sent, info = twilio_client.send_sms(
            to=r["to"], body=body, sender=config.SMS_SENDER_NUMBER)
        # Log before moving on: a send we failed to record would be re-sent.
        # In safe mode the text went to the TEST phone, not the patient, so it
        # must not count as sent — otherwise a rehearsal silently excludes real
        # people from the campaign.
        if not sent:
            status = f"failed: {info}"[:200]
        elif config.MARKETING_SAFE_MODE:
            status = "test-safe-mode (patient not contacted)"
        else:
            status = "sent"
        sent_log.log_send(r["pid"], r["name"], FLOW, "sms", ANCHOR, TEMPLATE,
                          status=status)
        ok, fail = (ok + 1, fail) if sent else (ok, fail + 1)
        if not sent:
            print(f"   FAILED {r['name']}: {info}")
        if i % 25 == 0:
            print(f"   {i}/{len(todo)}  ({ok} ok, {fail} failed)")
            sys.stdout.flush()
    print(f"\ndone: {ok} sent, {fail} failed")


if __name__ == "__main__":
    main()
