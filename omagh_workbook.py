"""Turn the Omagh catchment CSV into the marketing workbook.

Sheet 1 "Send List"   - one row per person, duplicate Cliniko records merged
Sheet 2 "All Records" - every matching record, unmerged, for audit
Sheet 3 "Summary"     - live COUNTIFS over the send list

Duplicate merge: same mobile, else same email, else same name + postcode.
The surviving row is the highest Patient ID (Cliniko ids are sequential, so
that is the most recent record and the most recent stated preference).
"""

import csv
import glob
import os
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

def _latest(pattern):
    """Newest matching export — the pull and the workbook build can land on
    different days, so today's date is not a safe assumption."""
    hits = sorted(glob.glob(os.path.expanduser(pattern)))
    if not hits:
        raise SystemExit(f"no file matching {pattern} — run omagh_list.py first")
    return hits[-1]


SRC = _latest("~/Downloads/omagh_catchment_*.csv")
STAMP = re.search(r"(\d{4}-\d{2}-\d{2})", SRC).group(1)
OUT = os.path.expanduser(f"~/Downloads/Omagh_Launch_List_{STAMP}.xlsx")

HEADERS = ["First name", "Name", "Email", "Mobile", "SMS-ready number", "Town", "Postcode",
           "Address", "Matched on", "Email marketing OK", "SMS marketing OK",
           "Consent note", "Duplicate records", "Patient ID"]

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)


def norm_mobile(m):
    d = re.sub(r"\D", "", m or "")
    if d.startswith("44"):
        d = "0" + d[2:]
    elif d.startswith("353"):
        d = "0" + d[3:]
    elif d and not d.startswith("0"):
        d = "0" + d
    return d if len(d) >= 10 else ""


def tidy_mobile(m):
    d = norm_mobile(m)
    if len(d) == 11 and d.startswith("07"):
        return f"{d[:5]} {d[5:]}"
    return (m or "").strip()


def sms_ready(m):
    """E.164 number, or "" if it is not a real mobile.

    Cliniko's "Mobile" type is free-typed, so the column holds 028 landlines
    and a few obvious keyboard-mashes alongside genuine numbers. Only UK 07
    and Irish 08 mobiles can actually receive a text.
    """
    d = norm_mobile(m)
    if len(d) == 11 and d.startswith("07"):
        return "+44" + d[1:]
    if len(d) == 10 and d.startswith("08"):
        return "+353" + d[1:]
    return ""


def key_for(r):
    mob = norm_mobile(r["Mobile"])
    if mob:
        return ("m", mob)
    email = r["Email"].strip().lower()
    if email:
        return ("e", email)
    return ("n", r["Name"].strip().lower(),
            re.sub(r"\s+", "", r["Postcode"]).upper())


def style(ws, n_rows, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = HEAD_FILL, HEAD_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{n_rows + 1}"
    for row in ws.iter_rows(min_row=2, max_row=n_rows + 1):
        for cell in row:
            cell.font = BODY_FONT
    # Phone numbers are text - a leading zero or + must survive.
    for row in ws.iter_rows(min_row=2, max_row=n_rows + 1,
                            min_col=4, max_col=5):
        for cell in row:
            cell.number_format = "@"


def write_sheet(ws, rows, widths):
    ws.append(HEADERS)
    for r in rows:
        ws.append([r.get(h, "") for h in HEADERS])
    style(ws, len(rows), widths)


def main():
    raw = list(csv.DictReader(open(SRC)))
    for r in raw:
        r["SMS-ready number"] = sms_ready(r["Mobile"])
        r["Mobile"] = tidy_mobile(r["Mobile"])
        r["Duplicate records"] = 1
        r["Consent note"] = ""
        # csv gives "True"/"False" text; the Summary COUNTIFS need real booleans.
        for flag in ("Email marketing OK", "SMS marketing OK"):
            r[flag] = r[flag].strip().lower() == "true"

    groups = {}
    for r in raw:
        groups.setdefault(key_for(r), []).append(r)

    merged = []
    for recs in groups.values():
        recs.sort(key=lambda r: int(r["Patient ID"]))
        best = dict(recs[-1])
        # Fill blanks from the older duplicates rather than losing contact details.
        for field in ("Email", "Mobile", "SMS-ready number", "Address",
                      "Postcode", "Town"):
            if not best.get(field):
                best[field] = next((r[field] for r in reversed(recs)
                                    if r.get(field)), "")
        # A duplicate record is created with the marketing flags off by default,
        # which is not the patient withdrawing consent. Consent given on any of
        # their records stands until they actually opt out.
        notes = []
        for flag, label in (("Email marketing OK", "email"),
                            ("SMS marketing OK", "SMS")):
            if not best[flag] and any(r[flag] for r in recs):
                best[flag] = True
                notes.append(f"{label} opt-in from an earlier duplicate record")
        best["Consent note"] = "; ".join(notes)
        best["Duplicate records"] = len(recs)
        merged.append(best)

    merged.sort(key=lambda r: (r["Town"] or "zz", r["Name"]))

    wb = Workbook()
    widths = [14, 24, 32, 15, 16, 15, 12, 40, 26, 11, 11, 38, 10, 20]
    write_sheet(wb.active, merged, widths)
    wb.active.title = "Send List"
    write_sheet(wb.create_sheet("All Records"), raw, widths)

    n = len(merged) + 1
    s = wb.create_sheet("Summary")
    s.column_dimensions["A"].width = 46
    s.column_dimensions["B"].width = 12
    s["A1"] = f"Omagh launch list - pulled from Cliniko {STAMP}"
    s["A1"].font = Font(name="Arial", size=12, bold=True)
    lines = [
        ("People on the send list (duplicates merged)",
         f"=COUNTA('Send List'!B2:B{n})"),
        ("Cliniko records behind them",
         f"=SUM('Send List'!M2:M{n})"),
        ("Have a first name for the merge",
         f"=COUNTIF('Send List'!A2:A{n},\"?*\")"),
        ("Have an email address", f"=COUNTIF('Send List'!C2:C{n},\"?*\")"),
        ("Have a mobile number", f"=COUNTIF('Send List'!D2:D{n},\"?*\")"),
        ("Of those, textable (real UK/IE mobile)",
         f"=COUNTIF('Send List'!E2:E{n},\"?*\")"),
        ("READY TO EMAIL (opted in + has email)",
         f"=COUNTIFS('Send List'!J2:J{n},TRUE,'Send List'!C2:C{n},\"?*\")"),
        ("READY TO TEXT (opted in + textable mobile)",
         f"=COUNTIFS('Send List'!K2:K{n},TRUE,'Send List'!E2:E{n},\"?*\")"),
        ("No email and no mobile at all",
         f"=COUNTIFS('Send List'!C2:C{n},\"\",'Send List'!D2:D{n},\"\")"),
    ]
    for i, (label, formula) in enumerate(lines, start=3):
        s.cell(row=i, column=1, value=label).font = BODY_FONT
        c = s.cell(row=i, column=2, value=formula)
        c.font = BODY_FONT
        c.alignment = Alignment(horizontal="right")
    note = s.cell(row=len(lines) + 4, column=1, value=(
        "Catchment: postcodes BT75-79, BT82, BT94, plus named Tyrone towns "
        "for records with no postcode. Source: Cliniko patient records, "
        "read-only pull. Consent flags are Cliniko's "
        "accepted_email_marketing / accepted_sms_marketing."))
    note.font = Font(name="Arial", size=9, italic=True)
    note.alignment = Alignment(wrap_text=True, vertical="top")
    s.merge_cells(start_row=len(lines) + 4, start_column=1,
                  end_row=len(lines) + 7, end_column=2)

    wb.save(OUT)
    print(f"send list: {len(merged)} people from {len(raw)} records")
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
